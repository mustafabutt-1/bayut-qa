#!/usr/bin/env python3
"""
Convert a designed test suite between JSON and a review spreadsheet.

Two export formats, same column layout:

    # cases -> csv  (flat, one row per case, Section repeated on every row --
    #                 the format for direct Testmo CSV import)
    python sheet_tools.py export --out "DPV Survey.csv" --stdin < cases.json
    python sheet_tools.py export cases.json --out "DPV Survey.csv"

    # cases -> xlsx  (section banner rows, per-platform status dropdowns, a Summary
    #                  sheet with live formulas -- the format for human review before
    #                  Testmo, e.g. in Notion / Google Sheets)
    python sheet_tools.py export --out "DPV Survey.xlsx" --app "Bayut UAE" --stdin < cases.json

    # xlsx -> JSON  (the upload skill does this itself; rarely run by hand)
    python sheet_tools.py import "DPV Survey.xlsx" --out cases.json

Format is inferred from --out's extension (.csv or .xlsx); override with --format if the
path doesn't carry one. CSV needs no dependency beyond the standard library -- xlsx needs
openpyxl, imported lazily so a pure-CSV run works even if it isn't installed.

The Description column holds a Gherkin scenario (Given/When/Then), which the Testmo
uploader parses into BDD steps.

IMPORTANT: the Testmo Case ID column is a record of what was uploaded, NOT a sync link.
Testmo has no update or delete API, so edits made in the sheet after upload CANNOT be
pushed back -- corrections to already-uploaded cases are a manual edit in the Testmo UI.

For Google Sheets: upload the .xlsx to Drive and open with Google Sheets (it converts on
open), or File > Import. Export back to .xlsx to re-import here.
"""

import argparse
import csv
import json
import sys
from pathlib import Path

FONT = "Arial"
STATUSES = ["Not Run", "Pass", "Fail", "Blocked", "N/A"]

COLUMNS = [
    ("#",                  6,   "index"),
    ("Section",            22,  "section"),
    ("Test Case",          58,  "title"),
    ("Description (BDD)",  66,  "description"),
    ("Expected Result",    58,  "expected"),
    ("Notes",              30,  "notes"),
    ("QA Status (Android)",18,  "status_android"),
    ("QA Status (iOS)",    18,  "status_ios"),
    ("Comments",           30,  "comments"),
    ("Testmo Case ID",     15,  "testmo_id"),
]

def _openpyxl():
    """Import openpyxl lazily -- only the xlsx path needs it; csv export/import don't."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        sys.exit("openpyxl is required for xlsx export/import:  pip install openpyxl")
    return locals()


# --------------------------------------------------------------------------- export csv

def export_csv(data, out_path):
    """Flat one-row-per-case CSV: Section repeated on every row, no banner rows, no
    styling -- this is the format Testmo's own CSV import expects, and matches the
    reference template the QA team supplied (see memory: testcase-csv-format)."""
    cases = data.get("cases", [])
    if not cases:
        sys.exit("No cases in input file")

    headers = [header for header, _, _ in COLUMNS]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for i, case in enumerate(cases, start=1):
            section = case.get("section", "General")
            row = {
                "index": i,
                "section": section,
                "title": case.get("title", ""),
                "description": case.get("description", ""),
                "expected": case.get("expected", ""),
                "notes": case.get("notes", ""),
                "status_android": "Not Run",
                "status_ios": "Not Run",
                "comments": case.get("comments", ""),
                "testmo_id": case.get("testmo_id", ""),
            }
            writer.writerow([row[key] for _, _, key in COLUMNS])
    return len(cases)


# --------------------------------------------------------------------------- export xlsx

def export_xlsx(data, out_path, app_name=None):
    ox = _openpyxl()
    Workbook = ox["Workbook"]
    Alignment, Border, Font = ox["Alignment"], ox["Border"], ox["Font"]
    PatternFill, Side = ox["PatternFill"], ox["Side"]
    get_column_letter = ox["get_column_letter"]
    DataValidation = ox["DataValidation"]

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
    BODY_FONT = Font(name=FONT, size=10)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    feature = data.get("feature", "Test Suite")
    cases = data.get("cases", [])
    if not cases:
        sys.exit("No cases in input file")

    wb = Workbook()

    # --- Test Cases sheet ---
    ws = wb.active
    ws.title = "Test Cases"

    for col, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    row = 2
    current_section = None
    for i, case in enumerate(cases, start=1):
        section = case.get("section", "General")

        # Section banner row, so the sheet reads like the team's existing checklists.
        if section != current_section:
            ws.cell(row=row, column=1, value=section)
            for col in range(1, len(COLUMNS) + 1):
                c = ws.cell(row=row, column=col)
                c.fill = SECTION_FILL
                c.font = Font(name=FONT, size=10, bold=True)
                c.border = BORDER
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=len(COLUMNS))
            current_section = section
            row += 1

        values = {
            "index": i,
            "section": section,
            "title": case.get("title", ""),
            "description": case.get("description", ""),
            "expected": case.get("expected", ""),
            "notes": case.get("notes", ""),
            "status_android": "Not Run",
            "status_ios": "Not Run",
            "comments": case.get("comments", ""),
            "testmo_id": case.get("testmo_id", ""),
        }

        for col, (_, _, key) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=row, column=col, value=values[key])
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(
                vertical="top",
                wrap_text=key in ("title", "description", "expected", "notes", "comments"),
                horizontal="center" if key in ("index", "status_android",
                                               "status_ios", "testmo_id") else "left",
            )
        row += 1

    last_row = row - 1

    dv = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"G2:H{last_row}")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    # --- Summary sheet (formulas, so it updates as QA fills the sheet in) ---
    s = wb.create_sheet("Summary")
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 18
    s.column_dimensions["C"].width = 18

    meta = [
        ("Feature", feature),
        ("App", app_name or data.get("app", "TODO — set the app")),
        ("Total test cases", len(cases)),
        ("Sections", len({c.get("section", "General") for c in cases})),
    ]
    r = 1
    for label, value in meta:
        s.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
        s.cell(row=r, column=2, value=value).font = BODY_FONT
        r += 1

    r += 1
    s.cell(row=r, column=1, value="Execution status").font = Font(name=FONT, size=11, bold=True)
    r += 1
    for col, header in enumerate(["", "Android", "iOS"], start=1):
        c = s.cell(row=r, column=col, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    header_row = r
    r += 1

    for status in STATUSES:
        s.cell(row=r, column=1, value=status).font = BODY_FONT
        s.cell(row=r, column=2,
               value=f'=COUNTIF(\'Test Cases\'!$G$2:$G${last_row},$A{r})').font = BODY_FONT
        s.cell(row=r, column=3,
               value=f'=COUNTIF(\'Test Cases\'!$H$2:$H${last_row},$A{r})').font = BODY_FONT
        r += 1

    s.cell(row=r, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
    for col in (2, 3):
        c = s.cell(row=r, column=col,
                   value=f'=SUM({get_column_letter(col)}{header_row + 1}:'
                         f'{get_column_letter(col)}{r - 1})')
        c.font = Font(name=FONT, size=10, bold=True)

    r += 2
    s.cell(row=r, column=1,
           value="Fill QA Status columns on the Test Cases sheet; these counts update "
                 "automatically.").font = Font(name=FONT, size=9, italic=True)
    r += 1
    s.cell(row=r, column=1,
           value="Description holds a Gherkin scenario (Given/When/Then) — the Testmo "
                 "uploader parses it into BDD steps.").font = Font(name=FONT, size=9, italic=True)
    r += 1
    s.cell(row=r, column=1,
           value="Testmo Case ID is filled after upload. Testmo has no update API — later sheet edits cannot be pushed back."
           ).font = Font(name=FONT, size=9, italic=True)

    wb.save(out_path)
    return len(cases), last_row


# --------------------------------------------------------------------------- import

def read_sheet(path):
    ox = _openpyxl()
    wb = ox["load_workbook"](path, data_only=True)
    if "Test Cases" not in wb.sheetnames:
        sys.exit(f"No 'Test Cases' sheet in {path}")
    ws = wb["Test Cases"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col

    # Accept either header spelling for the description column.
    desc_header = next((h for h in ("Description (BDD)", "Description") if h in headers), None)
    required = ["Test Case", "Expected Result"]
    missing = [h for h in required if h not in headers]
    if not desc_header:
        missing.append("Description (BDD)")
    if missing:
        sys.exit(f"Sheet is missing required columns: {', '.join(missing)}")

    def get(row, header):
        col = headers.get(header)
        if not col:
            return ""
        value = ws.cell(row=row, column=col).value
        return "" if value is None else str(value).strip()

    cases = []
    skipped_banners = 0
    for row in range(2, ws.max_row + 1):
        title = get(row, "Test Case")
        if not title:
            # Section banner rows have a value only in column A.
            if get(row, "Section") or ws.cell(row=row, column=1).value:
                skipped_banners += 1
            continue
        case = {
            "section": get(row, "Section") or "General",
            "title": title,
            "description": get(row, desc_header),
            "expected": get(row, "Expected Result"),
            "notes": get(row, "Notes"),
        }
        comments = get(row, "Comments")
        if comments:
            case["comments"] = comments
        testmo_id = get(row, "Testmo Case ID")
        if testmo_id:
            case["testmo_id"] = testmo_id
        cases.append(case)

    feature = Path(path).stem
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        for row in range(1, 10):
            if str(summary.cell(row=row, column=1).value or "").strip() == "Feature":
                feature = str(summary.cell(row=row, column=2).value or feature).strip()
                break

    return {"feature": feature, "cases": cases, "_skipped_banners": skipped_banners}


def import_sheet(path, out_path):
    """read_sheet + write the JSON to disk."""
    data = read_sheet(path)
    skipped = data.pop("_skipped_banners", 0)
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return len(data["cases"]), skipped


# --------------------------------------------------------------------------- cli

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="command", required=True)

    e = sub.add_parser("export", help="cases -> csv or xlsx")
    e.add_argument("input", nargs="?", help="JSON file of cases; omit to read stdin")
    e.add_argument("--stdin", action="store_true", help="read the cases JSON from stdin")
    e.add_argument("--out", required=True)
    e.add_argument("--app", help="App name for the Summary sheet (xlsx only)")
    e.add_argument("--format", choices=["csv", "xlsx"],
                   help="override format detection from --out's extension")

    i = sub.add_parser("import", help="xlsx -> JSON")
    i.add_argument("input")
    i.add_argument("--out", required=True)

    args = ap.parse_args()

    if args.command == "export":
        if args.input:
            with open(args.input) as f:
                data = json.load(f)
        else:
            data = json.load(sys.stdin)

        fmt = args.format or Path(args.out).suffix.lstrip(".").lower()
        if fmt not in ("csv", "xlsx"):
            sys.exit(f"can't infer format from --out {args.out!r}; pass --format csv|xlsx")

        if fmt == "csv":
            count = export_csv(data, args.out)
            print(f"Wrote {args.out}")
            print(f"  {count} cases, one row per case (Section repeated, no banner rows)")
            print("\nThis is the flat format for direct Testmo CSV import.")
        else:
            count, last_row = export_xlsx(data, args.out, args.app)
            print(f"Wrote {args.out}")
            print(f"  {count} cases, {last_row} rows including section banners")
            print(f"  Sheets: 'Test Cases' (edit here), 'Summary' (auto-counts)")
            print(f"\nFor Google Sheets: upload to Drive and open with Google Sheets.")
        print("No JSON copy is kept — the export is the deliverable.")
    else:
        count, banners = import_sheet(args.input, args.out)
        print(f"Wrote {args.out}")
        print(f"  {count} cases read, {banners} section banner rows skipped")
        print(f"\nNext: validate and upload with testmo_upload.py")


if __name__ == "__main__":
    main()
